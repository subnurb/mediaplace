"""Sync tool: playlist analysis and upload orchestration.

Flow:
  1. GET /api/sources/<id>/playlists/        → list playlists from a source
  2. POST /api/sync/                          → create SyncJob (returns job id)
  3. POST /api/sync/<id>/analyze/             → start background analysis thread
  4. GET  /api/sync/<id>/                     → poll: job + all tracks with statuses
  5. POST /api/sync/<id>/tracks/<tid>/upload/ → upload one NOT_FOUND/UNCERTAIN track
  6. POST /api/sync/<id>/push/                → push matched/confirmed tracks to target playlist
  7. GET  /api/sync/<id>/export/              → download match results as Excel
  8. GET  /api/sync/log/                      → history of all jobs with unsynced summaries
"""

import json as json_module
import logging
import os
import re
import tempfile
import threading
import traceback
from concurrent.futures import ThreadPoolExecutor, as_completed

logger = logging.getLogger(__name__)

import requests as http
import yt_dlp
from django.conf import settings
from django.db import close_old_connections
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt

from api.models import AudioFingerprint, SourceConnection, SyncJob, SyncTrack, TrackSource
from api.views import require_login
from music_matcher import classify_confidence, find_youtube_match
from video_creator import create_video
from youtube_uploader import upload_video_for_source

OUTPUT_DIR = settings.OUTPUT_DIR

# yt-dlp needs a JavaScript runtime for YouTube format extraction.
# Node.js ships at this path on macOS (Homebrew Intel); adjust if deployed elsewhere.
_NODE_PATH = "/usr/local/bin/node"
_YTDL_JS_OPTS: dict = (
    {"js_runtimes": {"node": {"path": _NODE_PATH}}}
    if os.path.exists(_NODE_PATH) else {}
)


# ── Source-type dispatchers ───────────────────────────────────────────────────

def _get_playlists_for_source(source):
    """Route to the correct playlist-fetching implementation by source type."""
    ST = SourceConnection.SourceType
    if source.source_type == ST.SOUNDCLOUD:
        from soundcloud_service import get_playlists
        return get_playlists(source)
    if source.source_type == ST.YOUTUBE_PUBLISH:
        from youtube_service import get_playlists
        return get_playlists(source)
    if source.source_type == ST.SPOTIFY:
        from spotify_service import get_playlists
        return get_playlists(source)
    raise NotImplementedError(
        f"Playlist browsing is not yet supported for {source.get_source_type_display()}."
    )


def _get_tracks_for_source(source, playlist_id):
    """Route to the correct track-fetching implementation by source type."""
    ST = SourceConnection.SourceType
    if source.source_type == ST.SOUNDCLOUD:
        from soundcloud_service import get_playlist_tracks
        return get_playlist_tracks(source, playlist_id)
    if source.source_type == ST.YOUTUBE_PUBLISH:
        from youtube_service import get_playlist_tracks
        return get_playlist_tracks(source, playlist_id)
    if source.source_type == ST.SPOTIFY:
        from spotify_service import get_playlist_tracks
        return get_playlist_tracks(source, playlist_id)
    raise NotImplementedError(
        f"Track fetching is not yet supported for {source.get_source_type_display()}."
    )


def _find_match_on_target(source_to, source_title, source_artist, source_duration_ms,
                          source_isrc=None, exclude_ids=None):
    """Route to the correct match-finding implementation by target source type.

    Returns (video_id, matched_title, confidence, alternatives).
    exclude_ids: list of previously-rejected target IDs to skip.
    """
    ST = SourceConnection.SourceType
    if source_to.source_type == ST.YOUTUBE_PUBLISH:
        return find_youtube_match(source_title, source_artist, source_duration_ms,
                                  source_isrc=source_isrc, exclude_ids=exclude_ids)
    if source_to.source_type == ST.SOUNDCLOUD:
        from soundcloud_service import find_soundcloud_match
        return find_soundcloud_match(source_to, source_title, source_artist, source_duration_ms,
                                     isrc=source_isrc, exclude_ids=exclude_ids)
    if source_to.source_type == ST.SPOTIFY:
        from spotify_service import find_spotify_match
        return find_spotify_match(source_to, source_title, source_artist, source_duration_ms,
                                  isrc=source_isrc, exclude_ids=exclude_ids)
    raise NotImplementedError(
        f"Track matching is not yet supported for {source_to.get_source_type_display()}."
    )


def _upload_to_target(source_to, track, tmp_dir):
    """Route to the correct upload implementation by target source type."""
    ST = SourceConnection.SourceType
    if source_to.source_type == ST.YOUTUBE_PUBLISH:
        # Check permanent audio cache first; fall back to a fresh download
        src_platform = track.job.source_from.source_type
        audio_path = _get_or_download_audio(
            src_platform, track.source_track_id, track.source_permalink_url
        )
        if not audio_path:
            audio_path = _download_audio(track.source_permalink_url, tmp_dir)
        artwork_path = _download_artwork(track.source_artwork_url, tmp_dir)
        video_path = os.path.join(tmp_dir, "output.mp4")
        create_video(artwork_path, audio_path, video_path, animation="none")
        return upload_video_for_source(
            source_to,
            video_path,
            title=track.source_title,
            description=f"Synced from {track.source_permalink_url}",
            tags=[track.source_artist] if track.source_artist else [],
            privacy="unlisted",
        )
    raise NotImplementedError(
        f"Upload is not yet supported for {source_to.get_source_type_display()}."
    )


# ── Confirmed-match cache ─────────────────────────────────────────────────────

def _check_confirmed_cache(src_platform: str, src_track_id: str, tgt_platform: str):
    """Return (video_id, title, 1.0, []) if this track has a user-confirmed match.

    Queries all previous SyncJobs for the same source platform + track_id where
    user_feedback='confirmed'. Called as a fast path before the expensive L1+L2
    search to reuse previously validated matches instantly.
    """
    hit = (
        SyncTrack.objects
        .filter(
            job__source_from__source_type=src_platform,
            source_track_id=src_track_id,
            job__source_to__source_type=tgt_platform,
            user_feedback="confirmed",
        )
        .exclude(target_video_id="")
        .order_by("-job__created_at")
        .values("target_video_id", "target_title")
        .first()
    )
    return (hit["target_video_id"], hit["target_title"], 1.0, []) if hit else None


# ── Source playlists ──────────────────────────────────────────────────────────

@require_login
def source_playlists(request, source_id):
    """GET /api/sources/<source_id>/playlists/"""
    try:
        source = request.user.sources.get(id=source_id)
    except SourceConnection.DoesNotExist:
        return JsonResponse({"error": "Source not found"}, status=404)

    try:
        playlists = _get_playlists_for_source(source)
        return JsonResponse({"playlists": playlists})
    except NotImplementedError as e:
        return JsonResponse({"error": str(e)}, status=400)
    except ValueError as e:
        return JsonResponse({"error": str(e)}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ── Sync job CRUD ─────────────────────────────────────────────────────────────

@csrf_exempt
@require_login
def sync_list_create(request):
    """GET /api/sync/ → list jobs   POST /api/sync/ → create job"""
    if request.method == "GET":
        jobs = request.user.sync_jobs.select_related("source_from", "source_to").all()
        return JsonResponse({"jobs": [j.to_dict() for j in jobs]})

    if request.method == "POST":
        try:
            data = json_module.loads(request.body)
            source_from = request.user.sources.get(id=data["source_from"])
            source_to = request.user.sources.get(id=data["source_to"])
        except (SourceConnection.DoesNotExist, KeyError) as e:
            return JsonResponse({"error": str(e)}, status=400)

        job = SyncJob.objects.create(
            user=request.user,
            source_from=source_from,
            source_to=source_to,
            playlist_id=data.get("playlist_id", ""),
            playlist_name=data.get("playlist_name", "Playlist"),
        )
        return JsonResponse(job.to_dict(), status=201)

    return JsonResponse({"error": "Method not allowed"}, status=405)


@require_login
def sync_detail(request, job_id):
    """GET /api/sync/<id>/ → job with all tracks"""
    try:
        job = request.user.sync_jobs.select_related("source_from", "source_to").get(id=job_id)
    except SyncJob.DoesNotExist:
        return JsonResponse({"error": "Job not found"}, status=404)

    return JsonResponse(job.to_dict(include_tracks=True))


# ── Audio cache ───────────────────────────────────────────────────────────────

def _get_or_download_audio(platform: str, track_id: str, url: str) -> str | None:
    """Return path to a full-quality cached audio file, downloading if not yet cached.

    Flow:
      1. Look up CachedAudio record for this TrackSource; return file_path if
         the file still exists on disk.
      2. Otherwise download at best available quality (converted to MP3 320 kbps)
         and persist a new CachedAudio record linked to the TrackSource.

    The cache directory is settings.AUDIO_CACHE_DIR (never a temp dir).
    Returns None if the download fails and no file is produced.
    """
    from api.models import CachedAudio

    ts = TrackSource.objects.filter(platform=platform, track_id=track_id).first()
    if ts:
        try:
            cache = ts.cached_audio
            if cache.file_path and os.path.exists(cache.file_path):
                logger.debug("[audio_cache] hit ts=%s  %s", ts.id, cache.file_path)
                return cache.file_path
            # Record exists but file is gone — fall through to re-download
        except CachedAudio.DoesNotExist:
            pass

    # Ensure cache directory exists
    cache_dir = str(settings.AUDIO_CACHE_DIR)
    os.makedirs(cache_dir, exist_ok=True)

    # Build a filesystem-safe filename from platform + track_id
    safe_id = re.sub(r"[^\w\-]", "_", str(track_id))[:80]
    name_base = f"{platform}_{safe_id}"

    ydl_opts = {
        **_YTDL_JS_OPTS,
        "format": "bestaudio/best",
        "outtmpl": os.path.join(cache_dir, f"{name_base}.%(ext)s"),
        "quiet": not settings.DEBUG,
        "no_warnings": not settings.DEBUG,
        "overwrites": True,
        "postprocessors": [{
            "key": "FFmpegExtractAudio",
            "preferredcodec": "mp3",
            "preferredquality": "320",
        }],
    }

    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            ydl.download([url])
    except yt_dlp.utils.DownloadError as exc:
        logger.warning("[audio_cache] download failed for %s: %s", url, exc)
        # Don't return yet — postprocessor failures sometimes still produce a file
    except Exception as exc:
        logger.warning("[audio_cache] unexpected error for %s: %s", url, exc, exc_info=True)

    # Find the downloaded/converted file (prefer .mp3, accept any extension)
    audio_path = None
    mp3_candidate = os.path.join(cache_dir, f"{name_base}.mp3")
    if os.path.exists(mp3_candidate) and os.path.getsize(mp3_candidate) > 0:
        audio_path = mp3_candidate
    else:
        for fname in sorted(os.listdir(cache_dir)):
            if fname.startswith(name_base + ".") and os.path.getsize(
                os.path.join(cache_dir, fname)
            ) > 0:
                audio_path = os.path.join(cache_dir, fname)
                break

    if not audio_path:
        return None

    # Persist to DB
    file_ext = os.path.splitext(audio_path)[1].lstrip(".")
    file_size = os.path.getsize(audio_path)
    if ts:
        CachedAudio.objects.update_or_create(
            track_source=ts,
            defaults={
                "file_path": audio_path,
                "file_format": file_ext,
                "file_size": file_size,
                "quality": "best",
            },
        )
        logger.debug(
            "[audio_cache] stored ts=%s  %s  (%d bytes)", ts.id, audio_path, file_size
        )
    return audio_path


# ── Level 3: audio feature comparison ────────────────────────────────────────

def _download_audio_snippet(url: str, dest_dir: str, prefix: str,
                             max_seconds: int = 30) -> str | None:
    """Download up to max_seconds of audio from URL using yt-dlp."""
    try:
        from yt_dlp.utils import download_range_func
    except ImportError:
        return None

    ydl_opts = {
        **_YTDL_JS_OPTS,
        "format": "bestaudio[ext=mp3]/bestaudio[protocol=https]/bestaudio[protocol=http]/bestaudio",
        "outtmpl": os.path.join(dest_dir, f"{prefix}.%(ext)s"),
        "quiet": not settings.DEBUG,
        "no_warnings": not settings.DEBUG,
        "overwrites": True,
        "download_ranges": download_range_func(None, [(0, max_seconds)]),
    }
    try:
        with yt_dlp.YoutubeDL(ydl_opts) as ydl:
            ydl.download([url])
    except yt_dlp.utils.DownloadError as exc:
        logger.warning("[snippet] download failed for %s: %s", url, exc)
        return None
    except Exception as exc:
        logger.warning("[snippet] unexpected error for %s: %s", url, exc, exc_info=True)
        return None

    for fname in sorted(os.listdir(dest_dir)):
        if fname.startswith(prefix + "."):
            return os.path.join(dest_dir, fname)
    return None


def _target_url(sync_track: SyncTrack, target_source_type: str) -> str:
    """Build the playable URL for the matched target track."""
    if target_source_type == SourceConnection.SourceType.YOUTUBE_PUBLISH:
        return f"https://www.youtube.com/watch?v={sync_track.target_video_id}"
    if target_source_type == SourceConnection.SourceType.SPOTIFY:
        return f"https://open.spotify.com/track/{sync_track.target_video_id}"
    return sync_track.target_video_id


def _get_or_build_fingerprint(platform: str, track_id: str,
                               audio_url: str, tmp_dir: str,
                               prefix: str) -> AudioFingerprint | None:
    """Return an AudioFingerprint for the given track, building one if needed.

    Analysis pipeline (all complementary):
      1. AcoustID + Chromaprint  → MBID, ISRCs, chromaprint string
      2. AcousticBrainz (or librosa fallback)  → BPM, key, mode
      3. ShazamIO  → Shazam ID, title, artist, album, genre, Spotify URI
      4. Local fingerprint (Dejavu-style)  → stored in LocalFingerprint table
    """
    from acoustid_service import lookup_mbid, get_acousticbrainz_features, get_mbid_isrcs
    from music_matcher import analyze_audio_features, MATCH_ALGO_VERSION
    from django.utils import timezone

    track_source = TrackSource.objects.filter(platform=platform, track_id=track_id).first()

    if track_source and track_source.fingerprint_id:
        fp = track_source.fingerprint
        # Return cached fingerprint only when it's current AND has audio features.
        # Missing bpm/key means AcousticBrainz returned nothing (shut down 2022);
        # fall through so the librosa fallback can fill them in.
        if not fp.is_stale() and (fp.bpm is not None or fp.key):
            if getattr(settings, "SHAZAM_ENABLED", False) and not fp.shazam_id:
                _run_shazam_sync(fp.id, track_source.id)
                fp.refresh_from_db()
            if getattr(settings, "LOCAL_FINGERPRINT_ENABLED", False):
                _run_local_fingerprint_sync(track_source.id, None)
            return fp

    # Use the persistent audio cache; fall back to a temp snippet if cache download fails
    audio_path = _get_or_download_audio(platform, track_id, audio_url)
    if not audio_path:
        audio_path = _download_audio_snippet(audio_url, tmp_dir, prefix)
    if not audio_path:
        return track_source.fingerprint if (track_source and track_source.fingerprint_id) else None

    mbid, acoustid_score, chromaprint_str = lookup_mbid(audio_path)

    now = timezone.now()

    if mbid:
        ab_feat = get_acousticbrainz_features(mbid)
        isrcs = get_mbid_isrcs(mbid)

        fp_data = {
            "chromaprint": chromaprint_str,
            "isrcs": isrcs,
            "source": "acoustid",
            "algo_version": MATCH_ALGO_VERSION,
        }
        if ab_feat.get("bpm"):
            fp_data["bpm"] = ab_feat["bpm"]
        if ab_feat.get("key"):
            fp_data["key"] = ab_feat["key"]
        if ab_feat.get("mode"):
            fp_data["mode"] = ab_feat["mode"]

        # AcousticBrainz was shut down in 2022 and often returns nothing.
        # Fall back to local librosa analysis for any missing features.
        if not fp_data.get("bpm") or not fp_data.get("key"):
            local_feat = analyze_audio_features(audio_path)
            if local_feat.get("tempo") and not fp_data.get("bpm"):
                fp_data["bpm"] = local_feat["tempo"]
            if local_feat.get("key") and not fp_data.get("key"):
                fp_data["key"] = local_feat["key"]
            if local_feat.get("mode") and not fp_data.get("mode"):
                fp_data["mode"] = local_feat["mode"]
            if fp_data.get("bpm") or fp_data.get("key"):
                fp_data["source"] = "acoustid+librosa"

        audio_fp, created = AudioFingerprint.objects.get_or_create(
            mbid=mbid, defaults=fp_data
        )
        if not created:
            for k, v in fp_data.items():
                setattr(audio_fp, k, v)
            audio_fp.save()

    else:
        local_feat = analyze_audio_features(audio_path)
        fp_data = {
            "chromaprint": chromaprint_str,
            "source": "librosa",
            "algo_version": MATCH_ALGO_VERSION,
        }
        if local_feat.get("tempo"):
            fp_data["bpm"] = local_feat["tempo"]
        if local_feat.get("key"):
            fp_data["key"] = local_feat["key"]
        if local_feat.get("mode"):
            fp_data["mode"] = local_feat["mode"]

        if track_source and track_source.fingerprint_id:
            audio_fp = track_source.fingerprint
            for k, v in fp_data.items():
                setattr(audio_fp, k, v)
            audio_fp.save()
        else:
            audio_fp = AudioFingerprint.objects.create(**fp_data)

    if track_source:
        track_source.fingerprint = audio_fp
        if not track_source.url:
            track_source.url = audio_url
        track_source.save(update_fields=["fingerprint", "url", "updated_at"])
    else:
        ts_obj, _ = TrackSource.objects.update_or_create(
            platform=platform, track_id=track_id,
            defaults={"fingerprint": audio_fp, "url": audio_url},
        )
        track_source = ts_obj

    # ── ShazamIO + local fingerprint — run synchronously so results feed into
    # _apply_audio_score for the Level-3 confidence calculation.
    # Shazam runs via subprocess (crash-safe); local FP via librosa.
    if getattr(settings, "SHAZAM_ENABLED", False) and not audio_fp.shazam_id:
        _run_shazam_sync(audio_fp.id, track_source.id if track_source else None)
        audio_fp.refresh_from_db()

    if getattr(settings, "LOCAL_FINGERPRINT_ENABLED", False) and track_source:
        _run_local_fingerprint_sync(track_source.id, audio_path)

    return audio_fp


def _run_shazam_sync(fp_id: int, ts_id: int | None) -> None:
    """Identify audio via ShazamIO and persist results (blocking).

    Runs shazam_service.py as a child subprocess so a segfault in the Rust
    extension only kills that child — the calling thread returns normally
    either way.  No-ops immediately if the fingerprint already has a shazam_id
    or no cached audio file is available.
    """
    import json
    import subprocess
    import sys
    from api.models import AudioFingerprint as _AF, TrackSource as _TS

    try:
        audio_path = None
        if ts_id:
            ts = _TS.objects.filter(id=ts_id).first()
            if ts:
                try:
                    ca = ts.cached_audio
                    if ca.file_path and os.path.exists(ca.file_path):
                        audio_path = ca.file_path
                except Exception:
                    pass
        if not audio_path:
            return

        service_script = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "shazam_service.py"
        )
        try:
            proc = subprocess.run(
                [sys.executable, service_script, audio_path],
                capture_output=True,
                timeout=60,
            )
        except subprocess.TimeoutExpired:
            logger.warning("[shazam] subprocess timed out for fp=%s", fp_id)
            return
        except Exception as exc:
            logger.warning("[shazam] subprocess error fp=%s: %s", fp_id, exc)
            return

        if proc.returncode != 0 or not proc.stdout.strip():
            return

        try:
            shazam_data = json.loads(proc.stdout.decode())
        except Exception:
            return

        if not shazam_data or not shazam_data.get("shazam_id"):
            return

        fp = _AF.objects.filter(id=fp_id).first()
        if not fp or fp.shazam_id:
            return

        update_fields = []
        for fp_field, data_key in [
            ("shazam_id", "shazam_id"),
            ("shazam_title", "title"),
            ("shazam_artist", "artist"),
            ("shazam_album", "album"),
            ("shazam_genre", "genre"),
            ("shazam_spotify_uri", "spotify_uri"),
            ("shazam_cover_url", "cover_url"),
        ]:
            value = shazam_data.get(data_key, "")
            if value and not getattr(fp, fp_field):
                setattr(fp, fp_field, value)
                update_fields.append(fp_field)

        if update_fields:
            update_fields.append("updated_at")
            fp.save(update_fields=update_fields)
            logger.debug("[shazam] fp=%s  id=%s  title=%s",
                         fp.id, fp.shazam_id, fp.shazam_title)
    except Exception as exc:
        logger.warning("[shazam] enrichment error fp=%s: %s", fp_id, exc)


def _schedule_shazam_enrichment(fp_id: int, ts_id: int | None) -> None:
    """Run _run_shazam_sync in a daemon thread (for non-L3 callers)."""
    import threading

    def _worker():
        from django.db import close_old_connections
        close_old_connections()
        try:
            _run_shazam_sync(fp_id, ts_id)
        finally:
            close_old_connections()

    threading.Thread(target=_worker, daemon=True).start()


def _run_local_fingerprint_sync(ts_id: int, audio_path: str | None) -> None:
    """Compute and store a Dejavu-style LocalFingerprint (blocking).

    No-ops if the fingerprint already exists for this TrackSource.
    """
    from api.models import LocalFingerprint, TrackSource as _TS

    try:
        if LocalFingerprint.objects.filter(track_source_id=ts_id).exists():
            return
        path = audio_path
        if not path:
            ts = _TS.objects.filter(id=ts_id).first()
            if ts:
                try:
                    ca = ts.cached_audio
                    if ca.file_path and os.path.exists(ca.file_path):
                        path = ca.file_path
                except Exception:
                    pass
        if not path:
            return
        import local_fingerprint_service
        local_fingerprint_service.store_fingerprint(ts_id, path)
    except Exception as exc:
        logger.warning("[local_fp] store error ts=%s: %s", ts_id, exc)


def _schedule_local_fingerprint(ts_id: int, audio_path: str | None) -> None:
    """Run _run_local_fingerprint_sync in a daemon thread (for non-L3 callers)."""
    import threading

    def _worker():
        from django.db import close_old_connections
        close_old_connections()
        try:
            _run_local_fingerprint_sync(ts_id, audio_path)
        finally:
            close_old_connections()

    threading.Thread(target=_worker, daemon=True).start()


def _apply_audio_score(conf: float, src_fp: AudioFingerprint | None,
                        tgt_fp: AudioFingerprint | None,
                        src_ts: TrackSource | None = None,
                        tgt_ts: TrackSource | None = None) -> float:
    """Compute adjusted confidence from fingerprint comparison.

    Priority (highest wins):
      1. Same MBID → 1.0
      2. Shared ISRC → 1.0
      3. Same Shazam ID → 1.0
      4. High local-fingerprint similarity (>= 0.15 Jaccard) → boost to 0.95
      5. BPM + key comparison → small boost/penalty
    """
    from music_matcher import bpm_match_boost

    src_mbid = src_fp.mbid if src_fp else ""
    tgt_mbid = tgt_fp.mbid if tgt_fp else ""

    if src_mbid and tgt_mbid:
        if src_mbid == tgt_mbid:
            return 1.0
        return max(0.0, conf - 0.15)

    if src_fp and tgt_fp:
        src_isrcs = set(src_fp.isrcs or [])
        tgt_isrcs = set(tgt_fp.isrcs or [])
        if src_isrcs and tgt_isrcs:
            if src_isrcs & tgt_isrcs:
                return 1.0
            return max(0.0, conf - 0.10)

        # Shazam ID match: strong signal that both sides are the same recording
        if src_fp.shazam_id and tgt_fp.shazam_id:
            if src_fp.shazam_id == tgt_fp.shazam_id:
                return 1.0

    # Local fingerprint (Dejavu-style) similarity
    if src_ts and tgt_ts:
        try:
            from api.models import LocalFingerprint
            import local_fingerprint_service

            src_lfp = LocalFingerprint.objects.filter(track_source=src_ts).first()
            tgt_lfp = LocalFingerprint.objects.filter(track_source=tgt_ts).first()
            if src_lfp and tgt_lfp:
                local_sim = local_fingerprint_service.similarity(
                    src_lfp.fingerprint_data, tgt_lfp.fingerprint_data
                )
                logger.debug(
                    "[local_fp] ts=%s vs ts=%s  jaccard=%.3f",
                    src_ts.id, tgt_ts.id, local_sim,
                )
                if local_sim >= 0.15:
                    # Very strong local fingerprint match: override to high confidence
                    return max(conf, 0.95)
                elif local_sim >= 0.05:
                    # Modest boost from partial local match
                    conf = min(1.0, conf + local_sim * 0.5)
        except Exception as exc:
            logger.debug("[local_fp] similarity check error: %s", exc)

    src_feat = src_fp.audio_features() if src_fp else {}
    tgt_feat = tgt_fp.audio_features() if tgt_fp else {}
    if not src_feat or not tgt_feat:
        return conf

    adjusted = bpm_match_boost(conf, src_feat, tgt_feat)
    src_key = src_feat.get("key")
    tgt_key = tgt_feat.get("key")
    if src_key and tgt_key:
        if src_key == tgt_key:
            mode_bonus = 0.03 if src_feat.get("mode") == tgt_feat.get("mode") else 0.01
            adjusted = min(1.0, adjusted + mode_bonus)
        else:
            adjusted = max(0.0, adjusted - 0.05)

    return adjusted


def _run_level3_one_track(sync_track_id: int, target_source_type: str, src_platform: str):
    """Worker: Level-3 fingerprint check for a single UNCERTAIN track."""
    from music_matcher import classify_confidence, THRESHOLD_MATCHED
    from django.utils import timezone

    close_old_connections()
    try:
        sync_track = SyncTrack.objects.get(id=sync_track_id)
        if not sync_track.target_video_id or not sync_track.source_permalink_url:
            return

        with tempfile.TemporaryDirectory() as tmp:
            src_fp = _get_or_build_fingerprint(
                platform=src_platform,
                track_id=sync_track.source_track_id,
                audio_url=sync_track.source_permalink_url,
                tmp_dir=tmp,
                prefix="src",
            )
            tgt_fp = _get_or_build_fingerprint(
                platform=target_source_type,
                track_id=sync_track.target_video_id,
                audio_url=_target_url(sync_track, target_source_type),
                tmp_dir=tmp,
                prefix="tgt",
            )

            if not src_fp and not tgt_fp:
                return

            # Fetch TrackSource objects for local fingerprint comparison
            src_ts = TrackSource.objects.filter(
                platform=src_platform, track_id=sync_track.source_track_id
            ).first()
            tgt_ts = TrackSource.objects.filter(
                platform=target_source_type, track_id=sync_track.target_video_id
            ).first()

            conf = sync_track.match_confidence or 0.0
            adjusted = round(_apply_audio_score(conf, src_fp, tgt_fp, src_ts, tgt_ts), 4)
            new_status = classify_confidence(adjusted)

            if adjusted != conf or new_status != sync_track.status:
                SyncTrack.objects.filter(id=sync_track_id).update(
                    match_confidence=adjusted,
                    status=new_status,
                )

            if adjusted >= THRESHOLD_MATCHED:
                now = timezone.now()
                for fp in (src_fp, tgt_fp):
                    if fp and fp.pk:
                        AudioFingerprint.objects.filter(pk=fp.pk).update(
                            match_count=fp.match_count + 1,
                            last_matched_at=now,
                        )
    except Exception:
        pass
    finally:
        close_old_connections()


def _merge_fingerprints(fp_id_a: int, fp_id_b: int) -> int:
    """Merge two AudioFingerprint records; return the surviving record's id.

    Winner selection: prefer MBID → more feature coverage → lower id (first seen).
    The loser's TrackSources are re-pointed to the winner.  Missing BPM / key /
    mode / ISRCs on the winner are filled in from the loser before deletion.
    """
    try:
        fp_a = AudioFingerprint.objects.get(id=fp_id_a)
        fp_b = AudioFingerprint.objects.get(id=fp_id_b)
    except AudioFingerprint.DoesNotExist:
        return fp_id_a  # already merged

    if fp_a.id == fp_b.id:
        return fp_a.id

    def _quality(fp):
        # Higher tuple = better candidate to keep
        return (bool(fp.mbid), bool(fp.bpm), bool(fp.key), bool(fp.isrcs), -fp.id)

    if _quality(fp_a) >= _quality(fp_b):
        winner, loser = fp_a, fp_b
    else:
        winner, loser = fp_b, fp_a

    # Fill missing values on winner from loser
    changed = []
    if not winner.mbid and loser.mbid:
        winner.mbid = loser.mbid
        changed.append("mbid")
    if winner.bpm is None and loser.bpm is not None:
        winner.bpm = loser.bpm
        changed.append("bpm")
    if not winner.key and loser.key:
        winner.key = loser.key
        changed.append("key")
    if not winner.mode and loser.mode:
        winner.mode = loser.mode
        changed.append("mode")
    if not winner.isrcs and loser.isrcs:
        winner.isrcs = loser.isrcs
        changed.append("isrcs")
    if not winner.chromaprint and loser.chromaprint:
        winner.chromaprint = loser.chromaprint
        changed.append("chromaprint")
    if changed:
        winner.save(update_fields=changed + ["updated_at"])

    # Re-point all TrackSources from loser to winner
    TrackSource.objects.filter(fingerprint=loser).update(fingerprint=winner)
    loser.delete()
    logger.debug("[fp_merge] merged fp=%s into winner fp=%s", loser.id, winner.id)
    return winner.id


def _persist_cross_platform_links(job_id: int) -> None:
    """Share AudioFingerprint records between matched TrackSource pairs.

    Called after analysis (and after user confirms a track) so the Library page
    can group cross-platform tracks immediately via fp.id grouping without
    waiting for a library re-sync.

    Logic: for every high-confidence or user-confirmed SyncTrack, if one side
    already has a fingerprint and the other doesn't, assign the same fingerprint
    to the other side.  If both already share a fingerprint — nothing to do.
    """
    from django.db.models import Q

    job = SyncJob.objects.select_related("source_from", "source_to").get(id=job_id)
    src_platform = job.source_from.source_type
    tgt_platform = job.source_to.source_type

    strong_matches = (
        SyncTrack.objects
        .filter(job=job)
        .filter(
            Q(status=SyncTrack.Status.MATCHED) |
            Q(status=SyncTrack.Status.UPLOADED) |
            Q(status=SyncTrack.Status.UNCERTAIN, user_feedback="confirmed")
        )
        .exclude(target_video_id="")
    )

    linked = 0
    for st in strong_matches:
        src_ts = (
            TrackSource.objects
            .filter(platform=src_platform, track_id=st.source_track_id)
            .only("id", "fingerprint_id")
            .first()
        )
        tgt_ts = (
            TrackSource.objects
            .filter(platform=tgt_platform, track_id=st.target_video_id)
            .only("id", "fingerprint_id")
            .first()
        )
        if not src_ts or not tgt_ts:
            continue
        if src_ts.fingerprint_id and not tgt_ts.fingerprint_id:
            TrackSource.objects.filter(id=tgt_ts.id).update(fingerprint_id=src_ts.fingerprint_id)
            linked += 1
        elif tgt_ts.fingerprint_id and not src_ts.fingerprint_id:
            TrackSource.objects.filter(id=src_ts.id).update(fingerprint_id=tgt_ts.fingerprint_id)
            linked += 1

    if linked:
        logger.info("[sync] job=%s linked %d cross-platform fingerprint pairs", job_id, linked)


def _run_level3_audio_check(job: SyncJob) -> None:
    """Level 3: fingerprint + compare source vs candidate for UNCERTAIN tracks.

    Runs in parallel using SYNC_ANALYSIS_PARALLELISM workers.
    Non-fatal: any per-track error keeps the original Level 1+2 result.
    """
    target_source_type = job.source_to.source_type
    src_platform = job.source_from.source_type

    uncertain_ids = list(
        SyncTrack.objects
        .filter(job=job, status=SyncTrack.Status.UNCERTAIN)
        .order_by("position")
        .values_list("id", flat=True)
    )
    if not uncertain_ids:
        return

    parallelism = getattr(settings, "SYNC_ANALYSIS_PARALLELISM", 5)

    with ThreadPoolExecutor(max_workers=parallelism) as executor:
        futs = [
            executor.submit(_run_level3_one_track, tid, target_source_type, src_platform)
            for tid in uncertain_ids
        ]
        for f in as_completed(futs):
            try:
                f.result()
            except Exception:
                pass


# ── Parallel match worker ─────────────────────────────────────────────────────

def _match_one_track(sync_track_id: int, job_id: int, isrc: str | None,
                     tgt_platform: str, src_platform: str) -> dict:
    """ThreadPoolExecutor worker: fetch the best match via API — NO DB writes.

    Reads are safe to run concurrently on SQLite; all writes are deferred to
    the orchestrator thread (see _run_analysis) so only one thread writes at a
    time, avoiding "database is locked" errors.

    Returns a result dict consumed by _apply_match_result().
    """
    close_old_connections()
    try:
        sync_track = SyncTrack.objects.get(id=sync_track_id)
        job = SyncJob.objects.select_related("source_to").get(id=job_id)

        # Fast path: same-platform sync (SC→SC, YT→YT).
        # The source track IS already on the target platform — no search needed.
        # Use the source track's own ID/URL directly with perfect confidence.
        if src_platform == tgt_platform:
            if tgt_platform == SourceConnection.SourceType.YOUTUBE_PUBLISH:
                video_id = sync_track.source_track_id
                tgt_url = f"https://www.youtube.com/watch?v={video_id}"
            else:
                # SoundCloud (and future platforms): stored "video_id" is the permalink URL
                video_id = sync_track.source_permalink_url or sync_track.source_track_id
                tgt_url = video_id
            logger.debug(
                "[sync_match] same-platform shortcut track=%s → %s",
                sync_track_id, video_id,
            )
            return {
                "ok": True,
                "sync_track_id": sync_track_id,
                "video_id": video_id,
                "matched_title": sync_track.source_title,
                "confidence": 1.0,
                "alternatives": [],
                "tgt_url": tgt_url,
                "tgt_platform": tgt_platform,
                "from_confirmed_cache": False,
            }

        # Fast path: previously confirmed by user (DB read only)
        cached = _check_confirmed_cache(src_platform, sync_track.source_track_id, tgt_platform)
        from_cache = False
        if cached:
            video_id, matched_title, confidence, alternatives = cached
            from_cache = True
            logger.debug(
                "[sync_match] cache hit track=%s src_id=%s → %s",
                sync_track_id, sync_track.source_track_id, video_id,
            )
        else:
            # Expensive network call — the reason we parallelize
            video_id, matched_title, confidence, alternatives = _find_match_on_target(
                job.source_to,
                sync_track.source_title,
                sync_track.source_artist,
                sync_track.source_duration_ms,
                source_isrc=isrc,
            )

        tgt_url = None
        if video_id:
            tgt_url = (
                f"https://www.youtube.com/watch?v={video_id}"
                if tgt_platform == SourceConnection.SourceType.YOUTUBE_PUBLISH
                else video_id
            )

        return {
            "ok": True,
            "sync_track_id": sync_track_id,
            "video_id": video_id or "",
            "matched_title": matched_title or "",
            "confidence": confidence,
            "alternatives": alternatives,
            "tgt_url": tgt_url,
            "tgt_platform": tgt_platform,
            "from_confirmed_cache": from_cache,
        }
    except Exception as err:
        return {
            "ok": False,
            "sync_track_id": sync_track_id,
            "error": str(err)[:500],
        }
    finally:
        close_old_connections()


def _apply_match_result(result: dict) -> None:
    """Persist a single match result returned by _match_one_track.

    Always called from the single orchestrator thread so SQLite never sees
    concurrent writers.
    """
    st_id = result["sync_track_id"]
    if not result.get("ok"):
        err_msg = result.get("error", "")
        logger.warning("[sync_match] track=%s error: %s", st_id, err_msg)
        SyncTrack.objects.filter(id=st_id).update(
            status=SyncTrack.Status.FAILED,
            error=err_msg[:500],
        )
        return

    video_id = result["video_id"]
    update_fields = dict(
        match_confidence=result["confidence"],
        target_video_id=video_id,
        target_title=result["matched_title"],
        status=classify_confidence(result["confidence"]) if video_id else SyncTrack.Status.NOT_FOUND,
        alternatives=result["alternatives"],
    )
    # When the match comes from the confirmed-match cache, carry the user's
    # prior approval forward so the track shows as "Confirmed" in the new job.
    if result.get("from_confirmed_cache") and video_id:
        update_fields["user_feedback"] = "confirmed"
    SyncTrack.objects.filter(id=st_id).update(**update_fields)

    if video_id and result.get("tgt_url"):
        TrackSource.objects.update_or_create(
            platform=result["tgt_platform"],
            track_id=video_id,
            defaults={"url": result["tgt_url"], "title": result["matched_title"]},
        )


# ── User match feedback ───────────────────────────────────────────────────────

@csrf_exempt
@require_login
def sync_confirm_track(request, job_id, track_id):
    """POST /api/sync/<id>/tracks/<tid>/confirm/"""
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        job = request.user.sync_jobs.select_related("source_from", "source_to").get(id=job_id)
        track = job.tracks.get(id=track_id)
    except (SyncJob.DoesNotExist, SyncTrack.DoesNotExist):
        return JsonResponse({"error": "Not found"}, status=404)

    track.user_feedback = "confirmed"
    track.save(update_fields=["user_feedback"])

    from django.utils import timezone
    now = timezone.now()
    src_platform = job.source_from.source_type
    tgt_platform = job.source_to.source_type

    # Update match_count on any existing fingerprints
    for platform, track_id_val in [
        (src_platform, track.source_track_id),
        (tgt_platform, track.target_video_id),
    ]:
        if not track_id_val:
            continue
        ts = TrackSource.objects.filter(platform=platform, track_id=track_id_val).first()
        if ts and ts.fingerprint_id:
            AudioFingerprint.objects.filter(pk=ts.fingerprint_id).update(
                match_count=ts.fingerprint.match_count + 1,
                last_matched_at=now,
            )

    # Share fingerprint between source and target so the Library groups them immediately
    if track.source_track_id and track.target_video_id:
        src_ts = (
            TrackSource.objects
            .filter(platform=src_platform, track_id=track.source_track_id)
            .only("id", "fingerprint_id")
            .first()
        )
        tgt_ts = (
            TrackSource.objects
            .filter(platform=tgt_platform, track_id=track.target_video_id)
            .only("id", "fingerprint_id")
            .first()
        )
        if src_ts and tgt_ts:
            if src_ts.fingerprint_id and not tgt_ts.fingerprint_id:
                TrackSource.objects.filter(id=tgt_ts.id).update(fingerprint_id=src_ts.fingerprint_id)
                logger.debug("[sync_confirm] fp link src=%s → tgt=%s", src_ts.id, tgt_ts.id)
            elif tgt_ts.fingerprint_id and not src_ts.fingerprint_id:
                TrackSource.objects.filter(id=src_ts.id).update(fingerprint_id=tgt_ts.fingerprint_id)
                logger.debug("[sync_confirm] fp link tgt=%s → src=%s", tgt_ts.id, src_ts.id)

    return JsonResponse(track.to_dict())


@csrf_exempt
@require_login
def sync_select_match(request, job_id, track_id):
    """POST /api/sync/<id>/tracks/<tid>/select/

    Manually select a candidate from the search_results list for a not_found track.
    Body: {"video_id": "<id or permalink>"}

    Sets the chosen candidate as the target, marks user_feedback='confirmed',
    and updates status to 'uncertain' (user confirmed, exact match not verified).
    """
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        data = json_module.loads(request.body)
        video_id = (data.get("video_id") or "").strip()
    except (json_module.JSONDecodeError, AttributeError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    if not video_id:
        return JsonResponse({"error": "video_id is required"}, status=400)

    try:
        job = request.user.sync_jobs.select_related("source_to").get(id=job_id)
        track = job.tracks.get(id=track_id)
    except (SyncJob.DoesNotExist, SyncTrack.DoesNotExist):
        return JsonResponse({"error": "Not found"}, status=404)

    # Find the chosen entry in stored alternatives (search_results)
    alternatives = track.alternatives or []
    chosen = next((a for a in alternatives if a.get("video_id") == video_id), None)
    if not chosen:
        chosen = {"video_id": video_id, "title": "", "artist": "", "confidence": 0.0}

    tgt_platform = job.source_to.source_type
    tgt_url = (
        f"https://www.youtube.com/watch?v={video_id}"
        if tgt_platform == SourceConnection.SourceType.YOUTUBE_PUBLISH
        else video_id
    )

    track.target_video_id = video_id
    track.target_title = chosen.get("title", "")
    track.match_confidence = chosen.get("confidence", 0.0)
    track.user_feedback = "confirmed"
    track.status = SyncTrack.Status.UNCERTAIN  # user-selected; not algorithm-verified
    track.alternatives = [a for a in alternatives if a.get("video_id") != video_id]

    track.save(update_fields=[
        "target_video_id", "target_title", "match_confidence",
        "status", "user_feedback", "alternatives",
    ])

    TrackSource.objects.update_or_create(
        platform=tgt_platform,
        track_id=video_id,
        defaults={"url": tgt_url, "title": track.target_title},
    )

    return JsonResponse(track.to_dict())


@csrf_exempt
@require_login
def sync_resolve_url(request, job_id, track_id):
    """POST /api/sync/<id>/tracks/<tid>/resolve-url/

    Body: {"url": "https://..."}
    Resolves a user-supplied URL from the target platform into track metadata
    so it can be added to the not_found picker list.
    Returns {video_id, title, artist}.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        data = json_module.loads(request.body)
        url = (data.get("url") or "").strip()
    except (json_module.JSONDecodeError, AttributeError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    if not url:
        return JsonResponse({"error": "url is required"}, status=400)

    try:
        job = request.user.sync_jobs.select_related("source_to").get(id=job_id)
    except SyncJob.DoesNotExist:
        return JsonResponse({"error": "Not found"}, status=404)

    tgt_type = job.source_to.source_type

    if tgt_type == SourceConnection.SourceType.YOUTUBE_PUBLISH:
        # Extract video ID from any YouTube URL form
        m = re.search(r'(?:[?&]v=|youtu\.be/)([A-Za-z0-9_-]{11})', url)
        if not m:
            return JsonResponse({"error": "Could not extract a YouTube video ID from this URL"}, status=400)
        video_id = m.group(1)
        title = ""
        artist = ""
        try:
            opts = {**_YTDL_JS_OPTS, "quiet": True, "no_warnings": True, "skip_download": True}
            with yt_dlp.YoutubeDL(opts) as ydl:
                info = ydl.extract_info(f"https://www.youtube.com/watch?v={video_id}", download=False)
                title = info.get("track") or info.get("title") or ""
                artist = info.get("artist") or info.get("channel") or info.get("uploader") or ""
        except Exception:
            pass
        return JsonResponse({"video_id": video_id, "title": title, "artist": artist})

    elif tgt_type == SourceConnection.SourceType.SOUNDCLOUD:
        try:
            from soundcloud_service import _get as sc_get, _extract_artist as sc_extract_artist
            track_data = sc_get(job.source_to, "/resolve", {"url": url})
            video_id = track_data.get("permalink_url") or url
            title = track_data.get("title") or ""
            artist = sc_extract_artist(track_data)
        except Exception as exc:
            return JsonResponse({"error": str(exc)}, status=400)
        return JsonResponse({"video_id": video_id, "title": title, "artist": artist})

    return JsonResponse({"error": "Unsupported target platform"}, status=400)


@csrf_exempt
@require_login
def sync_reject_track(request, job_id, track_id):
    """POST /api/sync/<id>/tracks/<tid>/reject/"""
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        job = request.user.sync_jobs.select_related("source_from", "source_to").get(id=job_id)
        track = job.tracks.get(id=track_id)
    except (SyncJob.DoesNotExist, SyncTrack.DoesNotExist):
        return JsonResponse({"error": "Not found"}, status=404)

    rejected_ids = list(track.rejected_target_ids or [])
    if track.target_video_id and track.target_video_id not in rejected_ids:
        rejected_ids.append(track.target_video_id)
    track.rejected_target_ids = rejected_ids
    track.user_feedback = ""

    remaining_alts = [
        a for a in (track.alternatives or [])
        if a.get("video_id") not in rejected_ids
    ]

    if remaining_alts:
        next_alt = remaining_alts[0]
        track.target_video_id = next_alt["video_id"]
        track.target_title = next_alt.get("title", "")
        track.match_confidence = next_alt.get("confidence", 0.0)
        track.status = classify_confidence(track.match_confidence)
        track.alternatives = remaining_alts[1:]
    else:
        try:
            video_id, matched_title, confidence, new_alts = _find_match_on_target(
                job.source_to,
                track.source_title,
                track.source_artist,
                track.source_duration_ms,
                exclude_ids=rejected_ids,
            )
            if video_id:
                track.target_video_id = video_id
                track.target_title = matched_title or ""
                track.match_confidence = confidence
                track.status = classify_confidence(confidence)
                track.alternatives = [
                    a for a in new_alts if a.get("video_id") not in rejected_ids
                ]
                tgt_platform = job.source_to.source_type
                tgt_url = (
                    f"https://www.youtube.com/watch?v={video_id}"
                    if tgt_platform == SourceConnection.SourceType.YOUTUBE_PUBLISH
                    else video_id
                )
                TrackSource.objects.update_or_create(
                    platform=tgt_platform,
                    track_id=video_id,
                    defaults={"url": tgt_url, "title": matched_title or ""},
                )
            else:
                track.target_video_id = ""
                track.target_title = ""
                track.match_confidence = 0.0
                track.status = SyncTrack.Status.NOT_FOUND
                track.alternatives = []
        except Exception:
            track.target_video_id = ""
            track.target_title = ""
            track.match_confidence = 0.0
            track.status = SyncTrack.Status.NOT_FOUND
            track.alternatives = []

    track.save(update_fields=[
        "target_video_id", "target_title", "match_confidence",
        "status", "user_feedback", "rejected_target_ids", "alternatives",
    ])
    return JsonResponse(track.to_dict())


# ── Analysis (background thread) ─────────────────────────────────────────────

def _run_analysis(job_id: int):
    """Background worker: fetch tracks from source, match against target in parallel."""
    from django.db import connection

    try:
        job = SyncJob.objects.select_related("source_from", "source_to").get(id=job_id)
        job.status = SyncJob.Status.ANALYZING
        job.save(update_fields=["status", "updated_at"])

        # 1. Fetch playlist tracks from the source platform
        tracks = _get_tracks_for_source(job.source_from, job.playlist_id)

        # 2. Create SyncTrack rows in PENDING state; upsert TrackSource for each source track
        SyncTrack.objects.filter(job=job).delete()
        isrc_map: dict[int, str] = {}  # SyncTrack.id → ISRC string
        src_platform = job.source_from.source_type
        tgt_platform = job.source_to.source_type

        for t in tracks:
            st = SyncTrack.objects.create(
                job=job,
                source_track_id=t["id"],
                source_title=t["title"],
                source_artist=t["artist"],
                source_duration_ms=t["duration_ms"],
                source_artwork_url=t["artwork_url"],
                source_permalink_url=t["permalink_url"],
                position=t["position"],
            )
            if t.get("isrc"):
                isrc_map[st.id] = t["isrc"]

            TrackSource.objects.update_or_create(
                platform=src_platform,
                track_id=t["id"],
                defaults={
                    "url": t.get("permalink_url", ""),
                    "title": t.get("title", ""),
                    "artist": t.get("artist", ""),
                    "duration_ms": t.get("duration_ms"),
                    "artwork_url": t.get("artwork_url", ""),
                },
            )

        # 3. Match each track against the target platform.
        #
        # Strategy for SQLite safety:
        #   - Workers run in parallel and do ONLY reads + network API calls
        #     (the expensive part). They return a result dict — no DB writes.
        #   - The orchestrator thread (here) writes each result serially via
        #     _apply_match_result(), so SQLite never sees concurrent writers.
        sync_track_ids = list(
            SyncTrack.objects.filter(job=job).order_by("position").values_list("id", flat=True)
        )
        parallelism = getattr(settings, "SYNC_ANALYSIS_PARALLELISM", 5)

        with ThreadPoolExecutor(max_workers=parallelism) as executor:
            futs = [
                executor.submit(
                    _match_one_track,
                    st_id, job.id, isrc_map.get(st_id), tgt_platform, src_platform,
                )
                for st_id in sync_track_ids
            ]
            for f in as_completed(futs):
                try:
                    _apply_match_result(f.result())
                except Exception as exc:
                    logger.error("[sync_analysis] job=%s _apply_match_result failed: %s",
                                 job_id, exc, exc_info=True)

        # 4. Level 3 — audio feature comparison for UNCERTAIN tracks (parallel)
        _run_level3_audio_check(job)

        # 5. Link cross-platform fingerprints for all strong/confirmed matches so the
        #    Library can group them immediately without a library re-sync.
        try:
            _persist_cross_platform_links(job.id)
        except Exception as exc:
            logger.warning("[sync_analysis] job=%s fp-link step failed: %s", job_id, exc)

        job.status = SyncJob.Status.READY
        job.save(update_fields=["status", "updated_at"])

    except Exception as exc:
        logger.error("[sync_analysis] job=%s FAILED: %s", job_id, exc)
        if settings.DEBUG:
            traceback.print_exc()
        try:
            SyncJob.objects.filter(id=job_id).update(status=SyncJob.Status.FAILED)
        except Exception:
            pass
    finally:
        connection.close()


@csrf_exempt
@require_login
def sync_analyze(request, job_id):
    """POST /api/sync/<id>/analyze/ — start background analysis."""
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        job = request.user.sync_jobs.get(id=job_id)
    except SyncJob.DoesNotExist:
        return JsonResponse({"error": "Job not found"}, status=404)

    if job.status in (SyncJob.Status.ANALYZING, SyncJob.Status.SYNCING):
        return JsonResponse({"error": "Job is already running"}, status=409)

    thread = threading.Thread(target=_run_analysis, args=(job_id,), daemon=True)
    thread.start()

    return JsonResponse({"status": "analyzing"})


# ── Push to playlist (background thread) ──────────────────────────────────────

def _playlist_existing_ids(source_to, playlist_id: str) -> set:
    """Return the set of track IDs already in the target playlist."""
    ST = SourceConnection.SourceType
    if source_to.source_type == ST.YOUTUBE_PUBLISH:
        from youtube_service import get_playlist_video_ids
        return get_playlist_video_ids(source_to, playlist_id)
    if source_to.source_type == ST.SOUNDCLOUD:
        from soundcloud_service import get_playlist_track_ids
        return get_playlist_track_ids(source_to, playlist_id)
    if source_to.source_type == ST.SPOTIFY:
        from spotify_service import get_playlist_track_ids
        return get_playlist_track_ids(source_to, playlist_id)
    return set()


def _add_to_playlist(source_to, playlist_id: str, track_ids: list) -> dict:
    """Add track_ids to the target playlist. Returns {added, already_existed, failed}."""
    ST = SourceConnection.SourceType
    existing = _playlist_existing_ids(source_to, playlist_id)

    added = 0
    already_existed = 0
    failed = 0

    if source_to.source_type == ST.YOUTUBE_PUBLISH:
        from youtube_service import add_video_to_playlist
        for vid in track_ids:
            if vid in existing:
                already_existed += 1
            else:
                ok = add_video_to_playlist(source_to, playlist_id, vid)
                if ok:
                    added += 1
                else:
                    failed += 1

    elif source_to.source_type == ST.SOUNDCLOUD:
        from soundcloud_service import add_tracks_to_playlist
        new_ids = [t for t in track_ids if t not in existing]
        already_existed = len(track_ids) - len(new_ids)
        if new_ids:
            ok = add_tracks_to_playlist(source_to, playlist_id, new_ids)
            if ok:
                added = len(new_ids)
            else:
                failed = len(new_ids)

    elif source_to.source_type == ST.SPOTIFY:
        from spotify_service import add_tracks_to_playlist
        new_ids = [t for t in track_ids if t not in existing]
        already_existed = len(track_ids) - len(new_ids)
        if new_ids:
            ok = add_tracks_to_playlist(source_to, playlist_id, new_ids)
            if ok:
                added = len(new_ids)
            else:
                failed = len(new_ids)

    return {"added": added, "already_existed": already_existed, "failed": failed}


def _run_push(job_id: int, target_playlist_id: str | None, new_playlist_name: str) -> None:
    """Background worker: push eligible tracks to the target playlist."""
    from django.db import connection
    from django.utils import timezone

    try:
        job = SyncJob.objects.select_related("source_to").get(id=job_id)
        job.status = SyncJob.Status.SYNCING
        job.save(update_fields=["status", "updated_at"])

        source_to = job.source_to

        # Create new playlist if requested
        if new_playlist_name and not target_playlist_id:
            ST = SourceConnection.SourceType
            if source_to.source_type == ST.YOUTUBE_PUBLISH:
                from youtube_service import create_playlist
                pl = create_playlist(source_to, new_playlist_name)
            elif source_to.source_type == ST.SOUNDCLOUD:
                from soundcloud_service import create_playlist
                pl = create_playlist(source_to, new_playlist_name)
            elif source_to.source_type == ST.SPOTIFY:
                from spotify_service import create_playlist
                pl = create_playlist(source_to, new_playlist_name)
            else:
                raise NotImplementedError("Playlist creation not supported for this platform.")
            target_playlist_id = pl["id"]
            playlist_name = pl["name"]
        else:
            playlist_name = new_playlist_name or target_playlist_id or ""

        # Collect eligible tracks:
        # - status = matched or uploaded → always add
        # - status = uncertain AND user_feedback = confirmed → user validated
        eligible_tracks = list(
            SyncTrack.objects.filter(job=job)
            .filter(
                status__in=[SyncTrack.Status.MATCHED, SyncTrack.Status.UPLOADED]
            )
            .exclude(target_video_id="")
        ) + list(
            SyncTrack.objects.filter(job=job)
            .filter(status=SyncTrack.Status.UNCERTAIN, user_feedback="confirmed")
            .exclude(target_video_id="")
        )

        track_ids = [t.target_video_id for t in eligible_tracks]
        eligible_by_id = {t.target_video_id: t for t in eligible_tracks}

        # Pre-fetch existing playlist contents for duplicate detection
        existing = _playlist_existing_ids(source_to, target_playlist_id)

        added = 0
        already_existed = 0
        failed = 0

        ST = SourceConnection.SourceType
        if source_to.source_type == ST.YOUTUBE_PUBLISH:
            from youtube_service import add_video_to_playlist
            for track_obj in eligible_tracks:
                vid = track_obj.target_video_id
                if vid in existing:
                    already_existed += 1
                    # Mark as pushed even if already existed
                    SyncTrack.objects.filter(id=track_obj.id).update(pushed_to_playlist=True)
                else:
                    ok = add_video_to_playlist(source_to, target_playlist_id, vid)
                    if ok:
                        added += 1
                        SyncTrack.objects.filter(id=track_obj.id).update(pushed_to_playlist=True)
                    else:
                        failed += 1

        elif source_to.source_type == ST.SOUNDCLOUD:
            from soundcloud_service import add_tracks_to_playlist, resolve_track_id

            # target_video_id for SoundCloud matches is a permalink URL (used for
            # display/audio download), but the playlist API needs numeric track IDs.
            # Resolve each URL to its numeric ID; use those for dedup and the PUT call.
            resolved_pairs = []  # [(numeric_id, track_obj)]
            for track_obj in eligible_tracks:
                tid = track_obj.target_video_id
                if tid.startswith("http"):
                    nid = resolve_track_id(source_to, tid)
                else:
                    nid = tid  # already numeric (future-proof)
                if nid:
                    resolved_pairs.append((nid, track_obj))
                else:
                    logger.warning(
                        "[sync_push] job=%s could not resolve SC track URL: %s", job_id, tid
                    )
                    failed += 1

            # Dedup against existing numeric IDs
            new_pairs = [(nid, t) for nid, t in resolved_pairs if nid not in existing]
            already_existed = len(resolved_pairs) - len(new_pairs)

            # Mark already-present tracks as pushed
            for nid, track_obj in resolved_pairs:
                if nid in existing:
                    SyncTrack.objects.filter(id=track_obj.id).update(pushed_to_playlist=True)

            if new_pairs:
                ok = add_tracks_to_playlist(
                    source_to, target_playlist_id,
                    [nid for nid, _ in new_pairs],
                )
                if ok:
                    added = len(new_pairs)
                    for _, track_obj in new_pairs:
                        SyncTrack.objects.filter(id=track_obj.id).update(pushed_to_playlist=True)
                else:
                    failed += len(new_pairs)

        elif source_to.source_type == ST.SPOTIFY:
            from spotify_service import add_tracks_to_playlist

            # target_video_id for Spotify matches is a bare track ID (22-char alphanumeric)
            new_ids = [t.target_video_id for t in eligible_tracks if t.target_video_id not in existing]
            already_existed = len(eligible_tracks) - len(new_ids)

            # Mark already-present tracks as pushed
            for track_obj in eligible_tracks:
                if track_obj.target_video_id in existing:
                    SyncTrack.objects.filter(id=track_obj.id).update(pushed_to_playlist=True)

            if new_ids:
                ok = add_tracks_to_playlist(source_to, target_playlist_id, new_ids)
                if ok:
                    added = len(new_ids)
                    new_id_set = set(new_ids)
                    for track_obj in eligible_tracks:
                        if track_obj.target_video_id in new_id_set:
                            SyncTrack.objects.filter(id=track_obj.id).update(pushed_to_playlist=True)
                else:
                    failed = len(new_ids)

        now = timezone.now()
        SyncJob.objects.filter(id=job_id).update(
            target_playlist_id=target_playlist_id,
            target_playlist_name=playlist_name,
            pushed_at=now,
            status=SyncJob.Status.DONE,
            updated_at=now,
        )

    except Exception as exc:
        logger.error("[sync_push] job=%s FAILED: %s", job_id, exc)
        if settings.DEBUG:
            traceback.print_exc()
        try:
            SyncJob.objects.filter(id=job_id).update(status=SyncJob.Status.FAILED)
        except Exception:
            pass
    finally:
        connection.close()


@csrf_exempt
@require_login
def sync_push(request, job_id):
    """POST /api/sync/<id>/push/

    Body: {"target_playlist_id": "PLxxx"|null, "new_playlist_name": "Name"|""}

    Eligibility: tracks with status=matched or status=uploaded are added automatically.
    Tracks with status=uncertain are added only if user_feedback='confirmed'.
    Skipped, not_found, failed tracks are ignored.

    Starts a background thread; job status transitions READY → SYNCING → DONE.
    Poll GET /api/sync/<id>/ to observe progress.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        job = request.user.sync_jobs.get(id=job_id)
    except SyncJob.DoesNotExist:
        return JsonResponse({"error": "Job not found"}, status=404)

    if job.status == SyncJob.Status.SYNCING:
        return JsonResponse({"error": "Push already in progress"}, status=409)

    try:
        data = json_module.loads(request.body)
    except Exception:
        data = {}

    target_playlist_id = data.get("target_playlist_id") or ""
    new_playlist_name = data.get("new_playlist_name") or ""

    if not target_playlist_id and not new_playlist_name:
        return JsonResponse(
            {"error": "Provide either target_playlist_id or new_playlist_name"}, status=400
        )

    thread = threading.Thread(
        target=_run_push,
        args=(job_id, target_playlist_id or None, new_playlist_name),
        daemon=True,
    )
    thread.start()

    return JsonResponse({"status": "syncing"})


# ── Excel export ──────────────────────────────────────────────────────────────

@require_login
def sync_export(request, job_id):
    """GET /api/sync/<id>/export/ — Download match results as .xlsx"""
    try:
        job = request.user.sync_jobs.select_related("source_from", "source_to").get(id=job_id)
    except SyncJob.DoesNotExist:
        return JsonResponse({"error": "Job not found"}, status=404)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return JsonResponse({"error": "openpyxl not installed"}, status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sync Results"

    # Header row
    headers = [
        "#", "Title", "Artist", "Duration", "Status", "Confidence",
        "Matched Title", "Match URL", "User Feedback", "Pushed",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    target_type = job.source_to.source_type

    for i, track in enumerate(job.tracks.all().order_by("position"), 1):
        if target_type == SourceConnection.SourceType.YOUTUBE_PUBLISH and track.target_video_id:
            target_url = f"https://www.youtube.com/watch?v={track.target_video_id}"
        else:
            target_url = track.target_video_id or ""

        dur = ""
        if track.source_duration_ms:
            s = track.source_duration_ms // 1000
            dur = f"{s // 60}:{s % 60:02d}"

        conf = ""
        if track.match_confidence is not None:
            conf = f"{round(track.match_confidence * 100)}%"

        ws.append([
            i,
            track.source_title,
            track.source_artist,
            dur,
            track.status,
            conf,
            track.target_title,
            target_url,
            track.user_feedback or "",
            "Yes" if track.pushed_to_playlist else "No",
        ])

    # Auto-size columns (approximate)
    col_widths = [5, 40, 25, 10, 12, 10, 40, 50, 15, 8]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Sanitize playlist name for filename
    safe_name = "".join(c for c in job.playlist_name if c.isalnum() or c in " _-")[:40].strip()

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="sync_{job_id}_{safe_name}.xlsx"'
    )
    wb.save(response)
    return response


# ── Sync history / log ────────────────────────────────────────────────────────

@require_login
def sync_log_view(request):
    """GET /api/sync/log/ — All jobs with per-status counts and unsynced track list.

    'Unsynced tracks' are defined as:
    - If the job has been pushed (pushed_at is set): tracks where pushed_to_playlist=False
      and status is not 'skipped' (these are tracks we couldn't or didn't add).
    - If the job has NOT been pushed yet: tracks with status in
      (not_found, uncertain, failed) that may need attention.
    """
    jobs = (
        request.user.sync_jobs
        .select_related("source_from", "source_to")
        .prefetch_related("tracks")
        .order_by("-created_at")
    )

    result = []
    for job in jobs:
        tracks = list(job.tracks.all())
        total = len(tracks)

        stats = {s: 0 for s in ["matched", "uncertain", "not_found", "uploaded",
                                  "uploading", "skipped", "failed", "pending"]}
        for t in tracks:
            stats[t.status] = stats.get(t.status, 0) + 1

        if job.pushed_at:
            # Job was pushed — show tracks not successfully added to playlist
            unsynced = [
                t.to_dict() for t in tracks
                if not t.pushed_to_playlist and t.status != SyncTrack.Status.SKIPPED
            ]
        else:
            # Job not pushed yet — show tracks that need attention
            attention_statuses = {
                SyncTrack.Status.NOT_FOUND,
                SyncTrack.Status.UNCERTAIN,
                SyncTrack.Status.FAILED,
            }
            unsynced = [t.to_dict() for t in tracks if t.status in attention_statuses]

        job_data = job.to_dict()
        job_data["stats"] = stats
        job_data["total_tracks"] = total
        job_data["unsynced_tracks"] = unsynced
        result.append(job_data)

    return JsonResponse({"jobs": result})


# ── Track upload (background thread) ─────────────────────────────────────────

def _run_track_upload(sync_track_id: int):
    """Background worker: download from SoundCloud + create video + upload to YouTube."""
    from django.db import connection

    try:
        track = SyncTrack.objects.select_related(
            "job__source_from", "job__source_to"
        ).get(id=sync_track_id)

        track.status = SyncTrack.Status.UPLOADING
        track.save(update_fields=["status"])

        with tempfile.TemporaryDirectory() as tmp:
            target_id = _upload_to_target(track.job.source_to, track, tmp)

        track.status = SyncTrack.Status.UPLOADED
        track.target_video_id = target_id or ""
        track.save(update_fields=["status", "target_video_id"])

    except Exception as err:
        try:
            SyncTrack.objects.filter(id=sync_track_id).update(
                status=SyncTrack.Status.FAILED,
                error=str(err)[:500],
            )
        except Exception:
            pass
    finally:
        connection.close()


def _download_audio(url: str, dest_dir: str) -> str:
    """Download best audio from a URL using yt-dlp. Returns the audio file path."""
    ydl_opts = {
        **_YTDL_JS_OPTS,
        "format": "bestaudio/best",
        "outtmpl": os.path.join(dest_dir, "audio.%(ext)s"),
        "quiet": True,
        "no_warnings": True,
        "postprocessors": [{
            "key": "FFmpegExtractAudio",
            "preferredcodec": "mp3",
            "preferredquality": "192",
        }],
    }
    with yt_dlp.YoutubeDL(ydl_opts) as ydl:
        ydl.download([url])

    for f in os.listdir(dest_dir):
        if f.startswith("audio."):
            return os.path.join(dest_dir, f)
    raise FileNotFoundError("yt-dlp download produced no audio file")


def _download_artwork(url: str, dest_dir: str) -> str:
    """Download artwork image. Falls back to a blank placeholder if unavailable."""
    if url:
        try:
            resp = http.get(url, timeout=10)
            resp.raise_for_status()
            ext = url.split(".")[-1].split("?")[0] or "jpg"
            path = os.path.join(dest_dir, f"artwork.{ext}")
            with open(path, "wb") as f:
                f.write(resp.content)
            return path
        except Exception:
            pass

    from PIL import Image
    path = os.path.join(dest_dir, "artwork.jpg")
    Image.new("RGB", (500, 500), (0, 0, 0)).save(path)
    return path


@csrf_exempt
@require_login
def sync_upload_track(request, job_id, track_id):
    """POST /api/sync/<id>/tracks/<tid>/upload/ — upload one track to target."""
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        job = request.user.sync_jobs.get(id=job_id)
        track = job.tracks.get(id=track_id)
    except (SyncJob.DoesNotExist, SyncTrack.DoesNotExist):
        return JsonResponse({"error": "Not found"}, status=404)

    if track.status == SyncTrack.Status.UPLOADING:
        return JsonResponse({"error": "Already uploading"}, status=409)

    thread = threading.Thread(target=_run_track_upload, args=(track_id,), daemon=True)
    thread.start()

    return JsonResponse({"status": "uploading"})


@csrf_exempt
@require_login
def sync_unconfirm_track(request, job_id, track_id):
    """POST /api/sync/<id>/tracks/<tid>/unconfirm/

    Remove the user's confirmation from a track (clears user_feedback).
    The match result is kept; only the confirmed flag is cleared.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        job = request.user.sync_jobs.get(id=job_id)
        track = job.tracks.get(id=track_id)
    except (SyncJob.DoesNotExist, SyncTrack.DoesNotExist):
        return JsonResponse({"error": "Not found"}, status=404)

    track.user_feedback = ""
    track.save(update_fields=["user_feedback"])
    return JsonResponse(track.to_dict())


@csrf_exempt
@require_login
def sync_confirm_all(request, job_id):
    """POST /api/sync/<id>/confirm-all/

    Confirm every track that has a match but hasn't been confirmed yet.
    Eligible: status in (matched, uncertain) with a non-empty target_video_id
    and user_feedback != 'confirmed'.
    Returns {"confirmed": N, "tracks": [...updated track dicts]}.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        job = request.user.sync_jobs.select_related("source_from", "source_to").get(id=job_id)
    except SyncJob.DoesNotExist:
        return JsonResponse({"error": "Job not found"}, status=404)

    from django.utils import timezone
    now = timezone.now()

    eligible = list(
        job.tracks
        .filter(status__in=[SyncTrack.Status.MATCHED, SyncTrack.Status.UNCERTAIN])
        .exclude(target_video_id="")
        .exclude(user_feedback="confirmed")
    )

    src_platform = job.source_from.source_type
    tgt_platform = job.source_to.source_type

    updated = []
    for track in eligible:
        track.user_feedback = "confirmed"
        track.save(update_fields=["user_feedback"])

        # Update fingerprint match counts
        for platform, track_id_val in [
            (src_platform, track.source_track_id),
            (tgt_platform, track.target_video_id),
        ]:
            if not track_id_val:
                continue
            ts = TrackSource.objects.filter(platform=platform, track_id=track_id_val).first()
            if ts and ts.fingerprint_id:
                AudioFingerprint.objects.filter(pk=ts.fingerprint_id).update(
                    match_count=ts.fingerprint.match_count + 1,
                    last_matched_at=now,
                )

        # Share fingerprint between source and target
        if track.source_track_id and track.target_video_id:
            src_ts = (
                TrackSource.objects
                .filter(platform=src_platform, track_id=track.source_track_id)
                .only("id", "fingerprint_id")
                .first()
            )
            tgt_ts = (
                TrackSource.objects
                .filter(platform=tgt_platform, track_id=track.target_video_id)
                .only("id", "fingerprint_id")
                .first()
            )
            if src_ts and tgt_ts:
                if src_ts.fingerprint_id and not tgt_ts.fingerprint_id:
                    TrackSource.objects.filter(id=tgt_ts.id).update(fingerprint_id=src_ts.fingerprint_id)
                elif tgt_ts.fingerprint_id and not src_ts.fingerprint_id:
                    TrackSource.objects.filter(id=src_ts.id).update(fingerprint_id=tgt_ts.fingerprint_id)

        updated.append(track.to_dict())

    return JsonResponse({"confirmed": len(updated), "tracks": updated})


@csrf_exempt
@require_login
def sync_skip_track(request, job_id, track_id):
    """POST /api/sync/<id>/tracks/<tid>/skip/"""
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        job = request.user.sync_jobs.get(id=job_id)
        track = job.tracks.get(id=track_id)
    except (SyncJob.DoesNotExist, SyncTrack.DoesNotExist):
        return JsonResponse({"error": "Not found"}, status=404)

    track.status = SyncTrack.Status.SKIPPED
    track.save(update_fields=["status"])
    return JsonResponse(track.to_dict())
